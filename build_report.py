#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
AI Tools Expense Report Builder
Generates professional Excel workbook with monthly and annual expense tracking.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.hyperlink import Hyperlink
import datetime, os, re
from pathlib import Path
from exchange_rate import fetch_usd_to_ils_rate
from manual_receipts_store import load_manual_receipts, receipt_identity, to_transaction_tuple

OUTPUT_PATH = (
    "AI_Tools_Expenses_2025_2026.xlsx"
    if os.environ.get("CI")
    else r"C:\Users\roita\מעקב הוצאות כלים\AI_Tools_Expenses_2025_2026.xlsx"
)
VENDOR_RULES_PATH = Path(__file__).with_name("vendor_rules.json")

# ── Color palette ──────────────────────────────────────────────────────────────
C_NAVY     = "1F3864"
C_BLUE     = "2E75B6"
C_LTBLUE   = "D6E4F0"
C_YELLOW   = "FFFF00"
C_WHITE    = "FFFFFF"
C_BLACK    = "000000"
C_DARKGRAY = "595959"
C_GREEN    = "70AD47"
C_ORANGE   = "ED7D31"
_EXCHANGE_RATE_CACHE = None

# Historical USD/ILS rates locked on the date of each charge.
# Source: Bank of Israel / fawazahmed0/currency-api.
# These ensure past transactions always display the rate from the charge date.
HISTORICAL_RATES = {
    "2025-07-04": 3.3365, "2025-07-09": 3.3483, "2025-07-10": 3.3092,
    "2025-07-11": 3.3151, "2025-07-16": 3.3656, "2025-07-20": 3.3579, "2025-07-30": 3.3663,
    "2025-08-01": 3.3944, "2025-08-02": 3.4127, "2025-08-04": 3.4102,
    "2025-08-05": 3.4176, "2025-08-09": 3.4338, "2025-08-10": 3.4332,
    "2025-08-11": 3.4215, "2025-08-16": 3.3771, "2025-08-30": 3.3421,
    "2025-09-01": 3.3362, "2025-09-05": 3.3481, "2025-09-10": 3.3434,
    "2025-09-16": 3.3506, "2025-09-24": 3.3429, "2025-09-30": 3.3013,
    "2025-10-02": 3.3163, "2025-10-16": 3.2858, "2025-10-24": 3.2902, "2025-10-30": 3.2531,
    "2025-11-09": 3.2599, "2025-11-11": 3.229,  "2025-11-13": 3.1955,
    "2025-11-16": 3.2342, "2025-11-24": 3.276,  "2025-11-30": 3.2617,
    "2025-12-01": 3.2583, "2025-12-03": 3.252,  "2025-12-08": 3.232,
    "2025-12-16": 3.2129, "2025-12-29": 3.1893, "2025-12-30": 3.1857,
    "2026-01-01": 3.1857, "2026-01-03": 3.1856, "2026-01-05": 3.1753,
    "2026-01-08": 3.1667, "2026-01-10": 3.1491, "2026-01-16": 3.1405,
    "2026-01-19": 3.1488, "2026-01-25": 3.1283, "2026-01-27": 3.1163, "2026-01-30": 3.0853,
    "2026-02-01": 3.1023, "2026-02-02": 3.1006, "2026-02-03": 3.099,
    "2026-02-10": 3.0831, "2026-02-12": 3.0771, "2026-02-15": 3.0863,
    "2026-02-16": 3.0886, "2026-02-24": 3.128,  "2026-02-25": 3.0992, "2026-02-28": 3.1354,
    "2026-03-01": 3.1283, "2026-03-02": 3.1204, "2026-03-03": 3.0772,
    "2026-03-10": 3.0952, "2026-03-12": 3.1308, "2026-03-15": 3.1436,
    "2026-03-16": 3.1433, "2026-03-19": 3.1112, "2026-03-24": 3.1177,
    "2026-03-25": 3.1249, "2026-03-26": 3.1184,
    "2026-04-01": 3.1433, "2026-04-02": 3.1385, "2026-04-03": 3.1329, "2026-04-05": 3.1414,
}


def get_current_usd_to_ils_rate():
    global _EXCHANGE_RATE_CACHE
    if _EXCHANGE_RATE_CACHE is None:
        _EXCHANGE_RATE_CACHE = fetch_usd_to_ils_rate()
    return _EXCHANGE_RATE_CACHE

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_BLACK, size=11, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def border_thin():
    s = Side(style="thin", color=C_DARKGRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color=C_NAVY)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def left_al():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── All transactions data ────────────────────────────────────────────────────
MANUAL_TRANSACTIONS = [
    # All amounts stored in ILS at the historical exchange rate on the day of charge.
    # Format: (date, tool, description, ils_amount, "ILS")
    # Rates source: Bank of Israel / fawazahmed0/currency-api historical data.
    ("2025-07-04", "Recraft",       "Basic – first month promo",              3.34, "ILS"),  # $1.00 @ 3.3365
    ("2025-07-09", "Ideogram AI",   "Ideogram Basic – annual",              241.08, "ILS"),  # $72.00 @ 3.3483
    ("2025-07-10", "Make",          "Core plan 10k ops/month",               35.04, "ILS"),  # $10.59 @ 3.3092
    ("2025-07-11", "Anthropic",     "Claude Pro",                            66.30, "ILS"),  # $20.00 @ 3.3151
    ("2025-07-20", "Runway ML",     "Standard (10% affiliate off)",          45.33, "ILS"),  # $13.50 @ 3.3579
    ("2025-07-30", "Replicate",     "Early card charge",                     33.66, "ILS"),  # $10.00 @ 3.3663
    ("2025-07-30", "Ideogram AI",   "Upgrade -> Plus Annual",               380.86, "ILS"),  # $113.14 @ 3.3663
    ("2025-08-01", "Google Workspace", "Business Plus – Jul 2025 (₪56.93)", 56.93, "ILS"),  # billed in ILS
    ("2025-08-02", "Runway ML",     "Credits 1,500",                         51.19, "ILS"),  # $15.00 @ 3.4127
    ("2025-08-04", "Recraft",       "Basic monthly",                         40.92, "ILS"),  # $12.00 @ 3.4102
    ("2025-08-05", "Replicate",     "Usage - July",                           3.08, "ILS"),  # $0.90 @ 3.4176
    ("2025-08-09", "Manus AI",      "Basic monthly",                         65.24, "ILS"),  # $19.00 @ 3.4338
    ("2025-08-10", "Make",          "Core plan 10k ops/month",               36.36, "ILS"),  # $10.59 @ 3.4332
    ("2025-08-11", "Anthropic",     "Claude Pro",                            68.43, "ILS"),  # $20.00 @ 3.4215
    ("2025-08-30", "Runway ML",     "Standard (25% off)",                    37.60, "ILS"),  # $11.25 @ 3.3421
    ("2025-08-30", "Eleven Labs",   "Starter monthly",                       16.71, "ILS"),  # $5.00 @ 3.3421
    ("2025-09-01", "Google Workspace", "Business Plus – Aug 2025 (₪56.93)", 56.93, "ILS"),  # billed in ILS
    ("2025-09-05", "Replicate",     "Usage - August",                         9.34, "ILS"),  # $2.79 @ 3.3481
    ("2025-09-10", "Make",          "Core plan 10k ops/month",               35.41, "ILS"),  # $10.59 @ 3.3434
    ("2025-09-24", "Anthropic",     "Claude Pro",                            66.86, "ILS"),  # $20.00 @ 3.3429
    ("2025-09-30", "Eleven Labs",   "Starter monthly",                       16.51, "ILS"),  # $5.00 @ 3.3013
    ("2025-10-02", "Google Workspace", "Business Plus – Sep 2025 (₪56.93)", 56.93, "ILS"),  # billed in ILS
    ("2025-10-24", "Anthropic",     "Claude Pro",                            65.80, "ILS"),  # $20.00 @ 3.2902
    ("2025-10-30", "Eleven Labs",   "Starter monthly",                       16.27, "ILS"),  # $5.00 @ 3.2531
    ("2025-11-09", "Recraft",       "Pro 1k credits",                        39.12, "ILS"),  # $12.00 @ 3.2599
    ("2025-11-11", "Astria",        "Credits x10 + 18% VAT",                38.10, "ILS"),  # $11.80 @ 3.229
    ("2025-11-11", "Runway ML",     "Standard monthly",                      48.44, "ILS"),  # $15.00 @ 3.229
    ("2025-11-13", "Hedra",         "Extra Small Credit Pack",               31.95, "ILS"),  # $10.00 @ 3.1955
    ("2025-11-24", "Anthropic",     "Claude Pro",                            65.52, "ILS"),  # $20.00 @ 3.276
    ("2025-11-30", "Eleven Labs",   "Starter monthly",                       16.31, "ILS"),  # $5.00 @ 3.2617
    ("2025-12-01", "Google Workspace", "Business Plus – Nov 2025 (₪75.90)", 75.90, "ILS"),  # billed in ILS
    ("2025-12-03", "Manychat",      "Pro monthly",                           48.78, "ILS"),  # $15.00 @ 3.252
    ("2025-12-03", "Meta (Ads)",    "Facebook Ads",                          31.61, "ILS"),  # $9.72 @ 3.252
    ("2025-12-08", "Meta (Ads)",    "Facebook Ads",                          96.96, "ILS"),  # $30.00 @ 3.232
    ("2025-12-16", "Meta (Ads)",    "Facebook Ads",                         106.03, "ILS"),  # $33.00 @ 3.2129
    ("2025-12-29", "Meta (Ads)",    "Facebook Ads",                         114.81, "ILS"),  # $36.00 @ 3.1893
    ("2025-12-30", "Eleven Labs",   "Starter monthly",                       15.93, "ILS"),  # $5.00 @ 3.1857
    ("2026-01-01", "Google Workspace", "Business Plus – Dec 2025 (₪75.90)", 75.90, "ILS"),  # billed in ILS
    ("2026-01-03", "Genspark",      "Plus Annual (happynewyear26 promo)",   445.73, "ILS"),  # $139.92 @ 3.1856
    ("2026-01-03", "Manychat",      "Pro monthly",                           47.78, "ILS"),  # $15.00 @ 3.1856
    ("2026-01-03", "Meta (Ads)",    "Facebook Ads",                          61.90, "ILS"),  # $19.43 @ 3.1856
    ("2026-01-05", "Genspark",      "Credits Pack",                          63.51, "ILS"),  # $20.00 @ 3.1753
    ("2026-01-08", "Genspark",      "Credits Pack",                          63.33, "ILS"),  # $20.00 @ 3.1667
    ("2026-01-10", "CapCut",        "Pro – Jan 2026 (₪49.90)",     49.90, "ILS"),  # billed in ILS
    ("2026-01-19", "Meta (Ads)",    "Facebook Ads",                         122.80, "ILS"),  # $39.00 @ 3.1488
    ("2026-01-25", "Lovable",       "Pro 1 monthly",                         78.21, "ILS"),  # $25.00 @ 3.1283
    ("2026-01-27", "Meta (Ads)",    "Facebook Ads",                         134.00, "ILS"),  # $43.00 @ 3.1163
    ("2026-01-30", "Timeless",      "Pro monthly (50% off)",                 44.74, "ILS"),  # $14.50 @ 3.0853
    ("2026-01-30", "Eleven Labs",   "Starter monthly",                       15.43, "ILS"),  # $5.00 @ 3.0853
    ("2026-02-01", "Lovable",       "Upgrade Pro1 -> Pro2 (prorated)",       77.56, "ILS"),  # $25.00 @ 3.1023
    ("2026-02-02", "Google Workspace", "Business Plus – Jan 2026 (₪75.90)", 75.90, "ILS"),  # billed in ILS
    ("2026-02-03", "Meta (Ads)",    "Facebook Ads",                          97.37, "ILS"),  # $31.42 @ 3.099
    ("2026-02-10", "CapCut",        "Pro – Feb 2026 (₪49.90)",     49.90, "ILS"),  # billed in ILS
    ("2026-02-12", "Meta (Ads)",    "Facebook Ads",                         144.62, "ILS"),  # $47.00 @ 3.0771
    ("2026-02-15", "Anthropic",     "Claude Pro",                            61.73, "ILS"),  # $20.00 @ 3.0863
    ("2026-02-16", "Anthropic",     "Credit purchase",                       15.44, "ILS"),  # $5.00 @ 3.0886
    ("2026-02-24", "Higgsfield",    "On-Demand credits 500",                 62.56, "ILS"),  # $20.00 @ 3.128
    ("2026-02-25", "Meta (Ads)",    "Facebook Ads",                         158.06, "ILS"),  # $51.00 @ 3.0992
    ("2026-02-25", "Lovable",       "Pro2 monthly",                         154.96, "ILS"),  # $50.00 @ 3.0992
    ("2026-02-28", "Eleven Labs",   "Starter monthly",                       15.68, "ILS"),  # $5.00 @ 3.1354
    ("2026-02-28", "Lovable",       "Cloud & AI Balance Top-up",             31.35, "ILS"),  # $10.00 @ 3.1354
    ("2026-03-01", "Lovable",       "Cloud & AI Balance Top-up",             31.28, "ILS"),  # $10.00 @ 3.1283
    ("2026-03-01", "Timeless",      "Pro monthly (50% off)",                 45.36, "ILS"),  # $14.50 @ 3.1283
    ("2026-03-01", "Google Workspace", "Google Workspace (₪75.9)",      75.90, "ILS"),  # billed in ILS
    ("2026-03-02", "Lovable",       "Cloud & AI Balance Top-up",             31.20, "ILS"),  # $10.00 @ 3.1204
    ("2026-03-02", "Google Workspace", "Business Plus – Feb 2026 (₪75.90)", 75.90, "ILS"),  # billed in ILS
    ("2026-03-03", "Meta (Ads)",    "Facebook Ads",                         123.43, "ILS"),  # $40.11 @ 3.0772
    ("2026-03-10", "CapCut",        "Pro – Mar 2026 (₪49.90)",     49.90, "ILS"),  # billed in ILS
    ("2026-03-12", "Meta (Ads)",    "Facebook Ads",                         159.67, "ILS"),  # $51.00 @ 3.1308
    ("2026-03-15", "Anthropic",     "Claude Pro",                            62.87, "ILS"),  # $20.00 @ 3.1436
    ("2026-03-18", "CapCut",        "Pro – Mar 2026 (₪49.90)",               49.90, "ILS"),  # 2nd account user4127233390216 (fwd)
    ("2026-03-19", "Anthropic",     "Credit purchase",                       31.11, "ILS"),  # $10.00 @ 3.1112
    ("2026-03-19", "Eleven Labs",   "Creator (first month 50% off)",         34.22, "ILS"),  # $11.00 @ 3.1112
    ("2026-03-25", "Lovable",       "Lite plan",                             15.62, "ILS"),  # $5.00 @ 3.1249
    ("2026-04-01", "Meta (Ads)",    "Facebook Ads",                         166.33, "ILS"),  # actual credit card charge
    ("2026-04-01", "Anthropic",     "Prepaid extra usage, Individual plan",  31.44, "ILS"),  # 2 × $5.00 @ 3.1433 (07:28 + 07:56 UTC)
    ("2026-04-01", "Google Workspace", "Business Plus – Mar 2026 (₪75.90)", 75.90, "ILS"),  # billed in ILS
    ("2026-04-03", "Meta (Ads)",    "Facebook Ads",                           7.05, "ILS"),  # $2.25 @ 3.1329
    ("2026-04-05", "IONOS",         "Instant Domain",                        62.83, "ILS"),  # $20.00 @ 3.1414
    ("2026-04-15", "Anthropic",     "Claude Pro",                            20.00),         # subscription
    ("2026-04-16", "Anthropic",     "Prepaid extra usage, Individual plan",  10.00),         # credit top-up
    ("2026-04-18", "CapCut",        "Pro – Apr 2026 (₪49.90)",               49.90, "ILS"),  # 2nd account user4127233390216 (fwd)
    ("2026-04-19", "Anthropic", "Prepaid extra usage, Individual plan", 5.00),
    ("2026-04-20", "Anthropic", "Auto recharge extra usage, Individual plan", 15.14),
    ("2026-04-20", "Anthropic", "Prepaid extra usage, Individual plan", 10.00),
    ("2026-04-21", "Anthropic", "Auto recharge extra usage, Individual plan", 10.15),
    ("2026-04-21", "Supabase", "Supabase", 25.00),
    ("2026-04-21", "Anthropic", "Auto recharge extra usage, Individual plan", 10.02),
    ("2026-04-21", "Anthropic", "Auto recharge extra usage, Individual plan", 10.50),
    ("2026-04-21", "Anthropic", "Auto recharge extra usage, Individual plan", 11.17),
    ("2026-04-22", "Anthropic", "Auto recharge extra usage, Individual plan", 15.57),
    ("2026-04-22", "Anthropic", "Auto recharge extra usage, Individual plan", 10.74),
    ("2026-04-23", "Anthropic", "Auto recharge extra usage, Individual plan", 10.16),
    ("2026-04-26", "Meta (Ads)", "Facebook Ads", 51.00),
    ("2026-04-26", "Lovable", "Lovable", 5.00),
    ("2026-05-01", "Google Cloud", "Google Cloud Platform & APIs (₪8.23)", 8.23, "ILS"),  # missed by Gmail scanner (no PDF)
]

TRACKING_START = datetime.date(2025, 7, 1)

HE_MONTH_NAMES = {
    1: "ינואר",
    2: "פברואר",
    3: "מרץ",
    4: "אפריל",
    5: "מאי",
    6: "יוני",
    7: "יולי",
    8: "אוגוסט",
    9: "ספטמבר",
    10: "אוקטובר",
    11: "נובמבר",
    12: "דצמבר",
}

EN_MONTH_NAMES = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

def load_vendor_rules():
    if not VENDOR_RULES_PATH.exists():
        return {}
    return json.loads(VENDOR_RULES_PATH.read_text(encoding="utf-8"))


VENDOR_RULES = load_vendor_rules()


def build_recurring_rules():
    rules = []
    for tool, config in VENDOR_RULES.items():
        if not config.get("subscription"):
            continue
        if config.get("billing_status") != "active":
            continue
        if not config.get("auto_add_to_sheet"):
            continue
        if not config.get("expected_amount") or not config.get("billing_day"):
            continue

        rules.append(
            {
                "tool": tool,
                "description": config.get("sheet_description") or tool,
                "amount": float(config["expected_amount"]),
                "currency": config.get("billing_currency", "USD"),
                "day": int(config["billing_day"]),
                "start_month": config.get("start_month", "2025-07"),
                "notes": f"Automatic recurring charge every month on the {config['billing_day']}.",
            }
        )
    return rules


RECURRING_RULES = build_recurring_rules()


def parse_ils_amount(description):
    match = re.search(r"₪\s*([0-9][0-9,]*(?:\.[0-9]+)?)", description)
    return float(match.group(1).replace(",", "")) if match else None


def normalize_transaction(txn):
    usd_rate, _, _ = get_current_usd_to_ils_rate()
    entry_source = txn[5] if len(txn) >= 6 else None

    if len(txn) >= 5:
        date, tool, description, original_amount, currency = txn[:5]
        original_amount = float(original_amount)
    else:
        date, tool, description, amount_usd = txn
        inferred_ils = parse_ils_amount(description)
        if inferred_ils is not None:
            original_amount = inferred_ils
            currency = "ILS"
        else:
            original_amount = float(amount_usd)
            currency = "USD"

    if currency == "ILS":
        amount_ils = round(original_amount, 2)
        amount_usd = float(txn[3]) if len(txn) == 4 else round(original_amount / usd_rate, 6)
    else:
        amount_usd = float(original_amount)
        # Use the locked historical rate for past charges; fall back to today's rate.
        hist_rate = HISTORICAL_RATES.get(date)
        effective_rate = hist_rate if hist_rate else usd_rate
        amount_ils = round(amount_usd * effective_rate, 2)

    return {
        "date": date,
        "tool": tool,
        "description": description,
        "currency": currency,
        "original_amount": round(original_amount, 2),
        "amount_usd": amount_usd,
        "amount_ils": amount_ils,
        "month_key": date[:7],
        "entry_source": entry_source,
    }


def first_day_of_month(d):
    return d.replace(day=1)


def add_months(d, months):
    year = d.year + (d.month - 1 + months) // 12
    month = (d.month - 1 + months) % 12 + 1
    return d.replace(year=year, month=month, day=1)


def build_months():
    months = []
    current = first_day_of_month(TRACKING_START)
    today = datetime.date.today()
    end = datetime.date(today.year, 12, 1)
    while current <= end:
        month_key = current.strftime("%Y-%m")
        months.append(
            (
                month_key,
                f"{HE_MONTH_NAMES[current.month]} {current.year}",
                f"{EN_MONTH_NAMES[current.month]} {current.year}",
            )
        )
        current = add_months(current, 1)
    return months


def build_recurring_transactions(months):
    month_keys = [month_key for month_key, _, _ in months]
    recurring = []
    for rule in RECURRING_RULES:
        for month_key in month_keys:
            if month_key < rule["start_month"]:
                continue
            if rule["currency"] == "ILS":
                recurring.append(
                    (
                        f"{month_key}-{rule['day']:02d}",
                        rule["tool"],
                        rule["description"],
                        rule["amount"],
                        "ILS",
                        "auto",
                    )
                )
            else:
                recurring.append(
                    (
                        f"{month_key}-{rule['day']:02d}",
                        rule["tool"],
                        rule["description"],
                        rule["amount"],
                        "USD",
                        "auto",
                    )
                )
    return recurring


def combine_transactions():
    manual = list(MANUAL_TRANSACTIONS)
    manual.extend(to_transaction_tuple(entry) for entry in load_manual_receipts())
    deduped_manual = []
    seen_manual = set()
    for txn in manual:
        txn_key = receipt_identity(normalize_transaction(txn))
        if txn_key in seen_manual:
            continue
        seen_manual.add(txn_key)
        deduped_manual.append(txn)

    recurring = build_recurring_transactions(MONTHS)
    manual_month_tool_desc = {(txn[0][:7], txn[1], txn[2]) for txn in deduped_manual}

    for txn in recurring:
        month_key = txn[0][:7]
        if (month_key, txn[1], txn[2]) in manual_month_tool_desc:
            continue
        deduped_manual.append(txn)

    return sorted(deduped_manual, key=lambda item: (item[0], item[1], item[2]))


MONTHS = build_months()
TRANSACTIONS = combine_transactions()
TRANSACTION_RECORDS = [normalize_transaction(txn) for txn in TRANSACTIONS]
ALL_TOOLS = sorted({transaction["tool"] for transaction in TRANSACTION_RECORDS})
REPORT_RANGE_HE = f"{MONTHS[0][1]} – {MONTHS[-1][1]}"
REPORT_RANGE_EN = f"{MONTHS[0][2]} – {MONTHS[-1][2]}"

# Sheet name for each month (short, safe for Excel tab)
MONTH_SHEET_NAMES = {mk: mname_he for mk, mname_he, _ in MONTHS}


def apply_header(ws, row, col, value, bg=C_NAVY, fg=C_WHITE, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill = fill(bg)
    cell.alignment = center()
    cell.border = border_thin()
    return cell


def apply_data(ws, row, col, value, bg=C_WHITE, fg=C_BLACK, bold=False,
               num_fmt=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=10, name="Arial")
    cell.fill = fill(bg)
    cell.border = border_thin()
    if align:
        cell.alignment = align
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def internal_link(ws, cell_addr, target_sheet, target_cell="A1", display=None):
    """Add a clickable internal hyperlink to another sheet."""
    cell = ws[cell_addr]
    if display is not None:
        cell.value = display
    cell.hyperlink = f"#{target_sheet}!{target_cell}"
    cell.font = Font(bold=True, color=C_WHITE, size=10, underline="single", name="Arial")


# ════════════════════════════════════════════════════════════════════════════════
# Sheet 1 – Settings
# ════════════════════════════════════════════════════════════════════════════════
def build_settings(wb):
    usd_rate, last_update, source = get_current_usd_to_ils_rate()
    ws = wb.create_sheet("\u05d4\u05d2\u05d3\u05e8\u05d5\u05ea")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = C_ORANGE

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "  \u05d4\u05d2\u05d3\u05e8\u05d5\u05ea \u05d5\u05e4\u05e8\u05de\u05d8\u05e8\u05d9\u05dd \u2013 AI & Business Tools Expenses"
    c.font = Font(bold=True, color=C_WHITE, size=16, name="Arial")
    c.fill = fill(C_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A3:B3")
    ws["A3"].value = "\u05e9\u05e2\u05e8 \u05d7\u05dc\u05d9\u05e4\u05d9\u05df"
    ws["A3"].font = Font(bold=True, color=C_WHITE, size=12, name="Arial")
    ws["A3"].fill = fill(C_BLUE)
    ws["A3"].alignment = left_al()

    ws["A4"].value = "\u05e9\u05e2\u05e8 USD \u2192 ILS (\u05de\u05ea\u05e2\u05d3\u05db\u05df \u05d0\u05d5\u05d8\u05d5\u05de\u05d8\u05d9\u05ea)"
    ws["A4"].font = Font(bold=False, size=10, name="Arial")
    ws["A4"].alignment = left_al()

    ws["B4"].value = usd_rate
    ws["B4"].font = Font(bold=True, color="0000FF", size=12, name="Arial")
    ws["B4"].fill = fill(C_YELLOW)
    ws["B4"].number_format = '#,##0.00" \u20aa/$"'
    ws["B4"].alignment = center()
    ws["B4"].border = border_medium()
    ws.row_dimensions[4].height = 22

    rate_context = f"\u2190 \u05de\u05e7\u05d5\u05e8: {source}"
    if last_update:
        rate_context += f" | \u05e2\u05d5\u05d3\u05db\u05df: {last_update[:10]}"
    ws["C4"].value = rate_context
    ws["C4"].font = Font(italic=True, size=9, color=C_DARKGRAY, name="Arial")

    ws["A6"].value = "\u05d4\u05e2\u05e8\u05d4:"
    ws["A6"].font = Font(bold=True, size=10, name="Arial")
    ws.merge_cells("B6:E8")
    ws["B6"].value = (
        "\u05db\u05dc \u05d4\u05e1\u05db\u05d5\u05de\u05d9\u05dd \u05d1\u05e9\u05e7\u05dc\u05d9\u05dd \u05de\u05d7\u05d5\u05e9\u05d1\u05d9\u05dd \u05e2\u05dc \u05e4\u05d9 \u05d4\u05e9\u05d9\u05e2\u05d5\u05e8 \u05e9\u05de\u05d5\u05d2\u05d3\u05e8 \u05d1\u05ea\u05d0 B4. "
        "\u05e0\u05d9\u05ea\u05df \u05dc\u05e9\u05e0\u05d5\u05ea \u05e9\u05d9\u05e2\u05d5\u05e8 \u05d6\u05d4 \u05dc\u05e4\u05d9 \u05d4\u05e9\u05d9\u05e2\u05d5\u05e8 \u05d4\u05d9\u05e1\u05d8\u05d5\u05e8\u05d9 \u05d4\u05e8\u05e6\u05d5\u05d9. "
        "\u05e9\u05d9\u05e2\u05d5\u05e8\u05d9 \u05d4\u05d4\u05de\u05e8\u05d4 \u05d1\u05e4\u05d5\u05e2\u05dc \u05e9\u05e9\u05d5\u05dc\u05de\u05d5 \u05e2\u05e9\u05d5\u05d9\u05d9\u05dd \u05dc\u05d4\u05d9\u05d5\u05ea \u05e9\u05d5\u05e0\u05d9\u05dd \u05de\u05e2\u05d8 \u05d1\u05e9\u05dc \u05e2\u05de\u05dc\u05d5\u05ea \u05d4\u05de\u05e8\u05d4."
    )
    ws["B6"].font = Font(italic=True, size=9, color=C_DARKGRAY, name="Arial")
    ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A10"].value = "חיובים קבועים:"
    ws["A10"].font = Font(bold=True, size=10, name="Arial")
    ws.merge_cells("B10:E12")
    ws["B10"].value = (
        'OpenAI / ChatGPT Plus – $20 בכל 16 לחודש. '
        "החיוב מתווסף אוטומטית לכל חודש בדוחות, בגליונות ובדשבורד, "
        "וכאשר תגיע חשבונית OpenAI במייל היא אמורה להיות מזוהה תחת אותו כלי."
    )
    ws["B10"].font = Font(italic=True, size=9, color=C_DARKGRAY, name="Arial")
    ws["B10"].alignment = Alignment(wrap_text=True, vertical="top")

    set_col_width(ws, 1, 35)
    set_col_width(ws, 2, 18)
    set_col_width(ws, 3, 45)

    return ws


# ════════════════════════════════════════════════════════════════════════════════
# Sheet 2 – All Transactions
# ════════════════════════════════════════════════════════════════════════════════
def build_transactions(wb):
    ws = wb.create_sheet("\u05db\u05dc \u05d4\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = C_BLUE
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"\u05db\u05dc \u05d4\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea \u2013 \u05db\u05dc\u05d9 AI  |  {REPORT_RANGE_HE}"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
    c.fill = fill(C_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 32

    headers = ["\u05ea\u05d0\u05e8\u05d9\u05da", "\u05db\u05dc\u05d9", "\u05ea\u05d9\u05d0\u05d5\u05e8", "\u05e1\u05db\u05d5\u05dd ($)", "\u05e1\u05db\u05d5\u05dd (\u20aa)", "\u05d7\u05d5\u05d3\u05e9"]
    for ci, h in enumerate(headers, 1):
        apply_header(ws, 2, ci, h, bg=C_BLUE)
    ws.row_dimensions[2].height = 22

    txns = sorted(TRANSACTION_RECORDS, key=lambda transaction: transaction["date"])

    for ri, transaction in enumerate(txns, 3):
        bg = C_LTBLUE if ri % 2 == 0 else C_WHITE
        month_name = next((m[1] for m in MONTHS if m[0] == transaction["month_key"]), transaction["month_key"])

        apply_data(ws, ri, 1, transaction["date"], bg=bg, align=center())
        apply_data(ws, ri, 2, transaction["tool"], bg=bg, align=left_al())
        apply_data(ws, ri, 3, transaction["description"], bg=bg, align=left_al())

        usd_cell = ws.cell(row=ri, column=4, value=transaction["amount_usd"])
        usd_cell.font = Font(size=10, name="Arial", color=C_BLACK)
        usd_cell.fill = fill(bg)
        usd_cell.border = border_thin()
        usd_cell.number_format = '[$$-en-US]#,##0.00'
        usd_cell.alignment = right()

        ils_cell = ws.cell(row=ri, column=5, value=transaction["amount_ils"])
        ils_cell.font = Font(size=10, name="Arial", color=C_BLACK)
        ils_cell.fill = fill(bg)
        ils_cell.border = border_thin()
        ils_cell.number_format = '#,##0.00" \u20aa"'
        ils_cell.alignment = right()

        apply_data(ws, ri, 6, month_name, bg=bg, align=center())

    last_row = len(txns) + 2
    total_row = last_row + 1
    ws.merge_cells(f"A{total_row}:C{total_row}")
    tc = ws.cell(row=total_row, column=1, value='\u05e1\u05d4"\u05db"')
    tc.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
    tc.fill = fill(C_NAVY)
    tc.alignment = center()
    tc.border = border_thin()

    for ci in [2, 3]:
        ws.cell(row=total_row, column=ci).fill = fill(C_NAVY)
        ws.cell(row=total_row, column=ci).border = border_thin()

    tot_usd = ws.cell(row=total_row, column=4, value=f"=SUM(D3:D{last_row})")
    tot_usd.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
    tot_usd.fill = fill(C_NAVY)
    tot_usd.border = border_medium()
    tot_usd.number_format = '[$$-en-US]#,##0.00'
    tot_usd.alignment = right()

    tot_ils = ws.cell(row=total_row, column=5, value=f"=SUM(E3:E{last_row})")
    tot_ils.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
    tot_ils.fill = fill(C_NAVY)
    tot_ils.border = border_medium()
    tot_ils.number_format = '#,##0.00" \u20aa"'
    tot_ils.alignment = right()

    ws.cell(row=total_row, column=6).fill = fill(C_NAVY)
    ws.cell(row=total_row, column=6).border = border_thin()

    widths = [14, 18, 40, 14, 14, 16]
    for ci, w in enumerate(widths, 1):
        set_col_width(ws, ci, w)

    return ws


# ════════════════════════════════════════════════════════════════════════════════
# Monthly Individual Sheets – one per month
# ════════════════════════════════════════════════════════════════════════════════
def build_month_sheets(wb):
    """Create one sheet per month with that month's transactions.
    Returns dict: month_key -> sheet_name"""
    rate_ref = "\u05d4\u05d2\u05d3\u05e8\u05d5\u05ea!$B$4"
    txns_sheet = "\u05db\u05dc \u05d4\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea"
    summary_sheet = "\u05e1\u05d9\u05db\u05d5\u05dd \u05d7\u05d5\u05d3\u05e9\u05d9"

    created = {}

    for mk, mname_he, mname_en in MONTHS:
        month_txns = sorted(
            [transaction for transaction in TRANSACTION_RECORDS if transaction["month_key"] == mk],
            key=lambda transaction: transaction["date"]
        )

        ws = wb.create_sheet(mname_he)
        ws.sheet_view.rightToLeft = True
        ws.sheet_properties.tabColor = "9DC3E6"
        ws.freeze_panes = "A3"
        created[mk] = mname_he

        # Title row
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = f"{mname_he}  |  {mname_en}"
        c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
        c.fill = fill(C_BLUE)
        c.alignment = center()
        ws.row_dimensions[1].height = 32

        # Back-link row
        back_cell = ws["A2"]
        back_cell.value = f"\u2190 \u05d7\u05d6\u05e8\u05d4 \u05dc\u05e1\u05d9\u05db\u05d5\u05dd \u05d7\u05d5\u05d3\u05e9\u05d9"
        back_cell.hyperlink = f"#'{summary_sheet}'!A1"
        back_cell.font = Font(bold=False, color="0000FF", size=9, underline="single", name="Arial")
        back_cell.fill = fill(C_LTBLUE)
        back_cell.alignment = left_al()
        back_cell.border = border_thin()
        for ci in range(2, 7):
            c = ws.cell(row=2, column=ci)
            c.fill = fill(C_LTBLUE)
            c.border = border_thin()
        ws.row_dimensions[2].height = 16

        # Column headers
        headers = ["\u05ea\u05d0\u05e8\u05d9\u05da", "\u05db\u05dc\u05d9", "\u05ea\u05d9\u05d0\u05d5\u05e8", "\u05e1\u05db\u05d5\u05dd ($)", "\u05e1\u05db\u05d5\u05dd (\u20aa)", ""]
        for ci, h in enumerate(headers, 1):
            apply_header(ws, 3, ci, h, bg=C_NAVY)
        ws.row_dimensions[3].height = 22

        if month_txns:
            for ri, transaction in enumerate(month_txns, 4):
                bg = C_LTBLUE if ri % 2 == 0 else C_WHITE
                apply_data(ws, ri, 1, transaction["date"], bg=bg, align=center())
                apply_data(ws, ri, 2, transaction["tool"], bg=bg, align=left_al())
                apply_data(ws, ri, 3, transaction["description"], bg=bg, align=left_al())

                c_usd = ws.cell(row=ri, column=4, value=transaction["amount_usd"])
                c_usd.font = Font(size=10, name="Arial")
                c_usd.fill = fill(bg)
                c_usd.border = border_thin()
                c_usd.number_format = '[$$-en-US]#,##0.00'
                c_usd.alignment = right()

                c_ils = ws.cell(row=ri, column=5, value=transaction["amount_ils"])
                c_ils.font = Font(size=10, name="Arial")
                c_ils.fill = fill(bg)
                c_ils.border = border_thin()
                c_ils.number_format = '#,##0.00" \u20aa"'
                c_ils.alignment = right()

                ws.cell(row=ri, column=6).fill = fill(bg)
                ws.cell(row=ri, column=6).border = border_thin()

            last_row = len(month_txns) + 3
            total_row = last_row + 1

            # Totals
            ws.merge_cells(f"A{total_row}:C{total_row}")
            tc = ws.cell(row=total_row, column=1, value='\u05e1\u05d4"\u05db \u05d7\u05d5\u05d3\u05e9"')
            tc.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
            tc.fill = fill(C_NAVY)
            tc.alignment = center()
            tc.border = border_medium()
            ws.cell(row=total_row, column=2).fill = fill(C_NAVY)
            ws.cell(row=total_row, column=2).border = border_medium()
            ws.cell(row=total_row, column=3).fill = fill(C_NAVY)
            ws.cell(row=total_row, column=3).border = border_medium()

            t_usd = ws.cell(row=total_row, column=4, value=f"=SUM(D4:D{last_row})")
            t_usd.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
            t_usd.fill = fill(C_NAVY)
            t_usd.border = border_medium()
            t_usd.number_format = '[$$-en-US]#,##0.00'
            t_usd.alignment = right()

            t_ils = ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{last_row})")
            t_ils.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
            t_ils.fill = fill(C_NAVY)
            t_ils.border = border_medium()
            t_ils.number_format = '#,##0.00" \u20aa"'
            t_ils.alignment = right()

            ws.cell(row=total_row, column=6).fill = fill(C_NAVY)
            ws.cell(row=total_row, column=6).border = border_medium()
            ws.row_dimensions[total_row].height = 24
        else:
            # Empty month
            ws.merge_cells("A4:F4")
            c = ws["A4"]
            c.value = "\u05d0\u05d9\u05df \u05d4\u05d5\u05e6\u05d0\u05d5\u05ea \u05d1\u05d7\u05d5\u05d3\u05e9 \u05d6\u05d4"
            c.font = Font(italic=True, size=10, color=C_DARKGRAY, name="Arial")
            c.fill = fill(C_LTBLUE)
            c.alignment = center()
            c.border = border_thin()

        # Column widths
        for ci, w in enumerate([14, 18, 40, 14, 14, 6], 1):
            set_col_width(ws, ci, w)

    return created


# ════════════════════════════════════════════════════════════════════════════════
# Sheet 3 – Monthly Summary Matrix (with SUMIFS + clickable month links)
# ════════════════════════════════════════════════════════════════════════════════
def build_monthly_summary(wb, month_sheet_names):
    ws = wb.create_sheet("\u05e1\u05d9\u05db\u05d5\u05dd \u05d7\u05d5\u05d3\u05e9\u05d9")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = C_GREEN
    ws.freeze_panes = "B4"

    rate_ref = "\u05d4\u05d2\u05d3\u05e8\u05d5\u05ea!$B$4"
    txns_sheet = "\u05db\u05dc \u05d4\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea"

    num_months = len(MONTHS)
    ws.merge_cells(f"A1:{get_column_letter(num_months * 2 + 2)}1")
    c = ws["A1"]
    c.value = f"\u05e1\u05d9\u05db\u05d5\u05dd \u05d7\u05d5\u05d3\u05e9\u05d9 \u05dc\u05e4\u05d9 \u05db\u05dc\u05d9  |  {REPORT_RANGE_HE}"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
    c.fill = fill(C_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 32

    # Instruction row
    ws.merge_cells(f"A2:{get_column_letter(num_months * 2 + 2)}2")
    c = ws["A2"]
    c.value = (
        "\u05dc\u05d7\u05e5 \u05e2\u05dc \u05e9\u05dd \u05d7\u05d5\u05d3\u05e9 \u05db\u05d3\u05d9 \u05dc\u05e6\u05e4\u05d9\u05d9\u05d4 "
        "\u05d1\u05d2\u05dc\u05d9\u05d5\u05df \u05d4\u05de\u05e4\u05d5\u05e8\u05d8 \u2193  |  OpenAI / ChatGPT Plus \u05de\u05ea\u05d5\u05d5\u05e1\u05e3 "
        "\u05d0\u05d5\u05d8\u05d5\u05de\u05d8\u05d9\u05ea \u05d1-16 \u05dc\u05db\u05dc \u05d7\u05d5\u05d3\u05e9"
    )
    c.font = Font(italic=True, size=9, color=C_DARKGRAY, name="Arial")
    c.fill = fill("F2F2F2")
    c.alignment = center()
    c.border = border_thin()
    ws.row_dimensions[2].height = 14

    # Header row: "Tool" then months (clickable)
    apply_header(ws, 3, 1, "\u05db\u05dc\u05d9", bg=C_NAVY)
    col = 2
    month_start_cols = {}
    for mi, (mk, mname_he, mname_en) in enumerate(MONTHS):
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        cell = ws.cell(row=3, column=col)
        cell.value = f"{mname_he}\n({mname_en})"
        cell.fill = fill(C_BLUE)
        cell.alignment = center()
        cell.border = border_thin()

        # Hyperlink to the individual month sheet (single quotes required for Hebrew sheet names)
        target_sheet = month_sheet_names.get(mk, mname_he)
        cell.hyperlink = f"#'{target_sheet}'!A1"
        cell.font = Font(bold=True, color=C_WHITE, size=9, underline="single", name="Arial")

        month_start_cols[mk] = col
        col += 2

    # Total columns
    total_usd_col = col
    total_ils_col = col + 1
    apply_header(ws, 3, total_usd_col, '\u05e1\u05d4"\u05db ($)', bg=C_NAVY)
    apply_header(ws, 3, total_ils_col, '\u05e1\u05d4"\u05db (\u20aa)', bg=C_NAVY)
    ws.row_dimensions[3].height = 42

    # Sub-header row for USD/ILS
    for mk, sc in month_start_cols.items():
        c1 = ws.cell(row=4, column=sc, value="$")
        c1.font = Font(bold=True, size=9, color=C_WHITE, name="Arial")
        c1.fill = fill(C_BLUE)
        c1.alignment = center()
        c1.border = border_thin()
        c2 = ws.cell(row=4, column=sc + 1, value="\u20aa")
        c2.font = Font(bold=True, size=9, color=C_WHITE, name="Arial")
        c2.fill = fill(C_BLUE)
        c2.alignment = center()
        c2.border = border_thin()

    ws.cell(row=4, column=1).fill = fill(C_NAVY)
    ws.cell(row=4, column=1).border = border_thin()
    for col_label, col_idx in [("$", total_usd_col), ("\u20aa", total_ils_col)]:
        c = ws.cell(row=4, column=col_idx, value=col_label)
        c.font = Font(bold=True, size=9, color=C_WHITE, name="Arial")
        c.fill = fill(C_NAVY)
        c.alignment = center()
        c.border = border_thin()
    ws.row_dimensions[4].height = 18

    # Data rows – SUMIFS pulling from "כל ההוצאות" so adding a row there auto-updates here
    data_start_row = 5
    for ti, tool in enumerate(ALL_TOOLS):
        row = data_start_row + ti
        bg = C_LTBLUE if ti % 2 == 0 else C_WHITE

        tc = ws.cell(row=row, column=1, value=tool)
        tc.font = Font(bold=True, size=10, name="Arial")
        tc.fill = fill(bg)
        tc.alignment = left_al()
        tc.border = border_thin()

        usd_sum_parts = []
        for mk, sc in month_start_cols.items():
            mname_he = MONTH_SHEET_NAMES[mk]
            # SUMIFS: sum col D in transactions sheet where col B = tool AND col F = month name
            sumif_formula = (
                f"=SUMIFS('{txns_sheet}'!$D:$D,"
                f"'{txns_sheet}'!$B:$B,\"{tool}\","
                f"'{txns_sheet}'!$F:$F,\"{mname_he}\")"
            )
            c_usd = ws.cell(row=row, column=sc, value=sumif_formula)
            c_usd.font = Font(size=10, name="Arial")
            c_usd.fill = fill(bg)
            c_usd.number_format = '[$$-en-US]#,##0.00;[$$-en-US](#,##0.00);"-"'
            c_usd.alignment = right()
            c_usd.border = border_thin()
            usd_sum_parts.append(f"{get_column_letter(sc)}{row}")

            c_ils = ws.cell(
                row=row,
                column=sc + 1,
                value=(
                    f"=SUMIFS('{txns_sheet}'!$E:$E,"
                    f"'{txns_sheet}'!$B:$B,\"{tool}\","
                    f"'{txns_sheet}'!$F:$F,\"{mname_he}\")"
                ),
            )
            c_ils.font = Font(size=10, name="Arial")
            c_ils.fill = fill(bg)
            c_ils.number_format = '#,##0.00" \u20aa";(#,##0.00" \u20aa");"-"'
            c_ils.alignment = right()
            c_ils.border = border_thin()

        # Row totals
        t_usd = ws.cell(row=row, column=total_usd_col,
                        value=f"=SUM({','.join(usd_sum_parts)})")
        t_usd.font = Font(bold=True, size=10, name="Arial")
        t_usd.fill = fill(C_LTBLUE)
        t_usd.number_format = '[$$-en-US]#,##0.00;[$$-en-US](#,##0.00);"-"'
        t_usd.alignment = right()
        t_usd.border = border_thin()

        t_ils = ws.cell(row=row, column=total_ils_col)
        t_ils.value = f"=SUM({','.join(get_column_letter(sc + 1) + str(row) for _, sc in month_start_cols.items())})"
        t_ils.font = Font(bold=True, size=10, name="Arial")
        t_ils.fill = fill(C_LTBLUE)
        t_ils.number_format = '#,##0.00" \u20aa";(#,##0.00" \u20aa");"-"'
        t_ils.alignment = right()
        t_ils.border = border_thin()

    # Grand total row
    tot_row = data_start_row + len(ALL_TOOLS)
    tc = ws.cell(row=tot_row, column=1, value='\u05e1\u05d4"\u05db \u05d7\u05d5\u05d3\u05e9\u05d9')
    tc.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
    tc.fill = fill(C_NAVY)
    tc.alignment = center()
    tc.border = border_medium()

    for mk, sc in month_start_cols.items():
        t_usd = ws.cell(row=tot_row, column=sc,
            value=f"=SUM({get_column_letter(sc)}{data_start_row}:{get_column_letter(sc)}{tot_row-1})")
        t_usd.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
        t_usd.fill = fill(C_NAVY)
        t_usd.number_format = '[$$-en-US]#,##0.00'
        t_usd.alignment = right()
        t_usd.border = border_medium()

        t_ils = ws.cell(row=tot_row, column=sc + 1,
            value=f"=SUM({get_column_letter(sc+1)}{data_start_row}:{get_column_letter(sc+1)}{tot_row-1})")
        t_ils.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
        t_ils.fill = fill(C_NAVY)
        t_ils.number_format = '#,##0.00" \u20aa"'
        t_ils.alignment = right()
        t_ils.border = border_medium()

    gt_usd = ws.cell(row=tot_row, column=total_usd_col,
        value=f"=SUM({get_column_letter(total_usd_col)}{data_start_row}:{get_column_letter(total_usd_col)}{tot_row-1})")
    gt_usd.font = Font(bold=True, color=C_WHITE, size=12, name="Arial")
    gt_usd.fill = fill(C_NAVY)
    gt_usd.number_format = '[$$-en-US]#,##0.00'
    gt_usd.alignment = right()
    gt_usd.border = border_medium()

    gt_ils = ws.cell(row=tot_row, column=total_ils_col,
        value=f"=SUM({get_column_letter(total_ils_col)}{data_start_row}:{get_column_letter(total_ils_col)}{tot_row-1})")
    gt_ils.font = Font(bold=True, color=C_WHITE, size=12, name="Arial")
    gt_ils.fill = fill(C_NAVY)
    gt_ils.number_format = '#,##0.00" \u20aa"'
    gt_ils.alignment = right()
    gt_ils.border = border_medium()

    ws.row_dimensions[tot_row].height = 24

    # Column widths
    set_col_width(ws, 1, 20)
    for mk, sc in month_start_cols.items():
        set_col_width(ws, sc, 11)
        set_col_width(ws, sc + 1, 13)
    set_col_width(ws, total_usd_col, 12)
    set_col_width(ws, total_ils_col, 14)

    return ws


# ════════════════════════════════════════════════════════════════════════════════
# Sheet 4 – Annual Report
# ════════════════════════════════════════════════════════════════════════════════
def build_annual_report(wb, month_sheet_names):
    ws = wb.create_sheet("\u05d3\u05d5\u05d7 \u05e9\u05e0\u05ea\u05d9")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = "ED7D31"

    rate_ref = "\u05d4\u05d2\u05d3\u05e8\u05d5\u05ea!$B$4"
    txns_sheet = "\u05db\u05dc \u05d4\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "\u05d3\u05d5\u05d7 \u05e9\u05e0\u05ea\u05d9 \u2013 AI & Business Tools Expenses"
    c.font = Font(bold=True, color=C_WHITE, size=16, name="Arial")
    c.fill = fill(C_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"\u05e1\u05d9\u05db\u05d5\u05dd \u05d4\u05d5\u05e6\u05d0\u05d5\u05ea \u05dc\u05e4\u05d9 \u05d7\u05d5\u05d3\u05e9 \u2013 {REPORT_RANGE_HE}"
    c.font = Font(bold=False, color=C_WHITE, size=11, italic=True, name="Arial")
    c.fill = fill(C_BLUE)
    c.alignment = center()
    ws.row_dimensions[2].height = 20

    hdrs = ["\u05d7\u05d5\u05d3\u05e9", "\u05ea\u05d0\u05e8\u05d9\u05da", '\u05d4\u05d5\u05e6\u05d0\u05d4 ($)', '\u05d4\u05d5\u05e6\u05d0\u05d4 (\u20aa)', '% \u05de\u05e1\u05d4"\u05db', "\u05d4\u05e2\u05e8\u05d5\u05ea"]
    for ci, h in enumerate(hdrs, 1):
        apply_header(ws, 3, ci, h, bg=C_NAVY)
    ws.row_dimensions[3].height = 22

    row = 4
    for mi, (mk, mname_he, mname_en) in enumerate(MONTHS):
        bg = C_LTBLUE if mi % 2 == 0 else C_WHITE

        # Month name cell: clickable link to individual month sheet
        month_cell = ws.cell(row=row, column=1, value=mname_he)
        month_cell.font = Font(bold=True, color="0000FF", size=10, underline="single", name="Arial")
        month_cell.fill = fill(bg)
        month_cell.border = border_thin()
        month_cell.alignment = left_al()
        target_sheet = month_sheet_names.get(mk, mname_he)
        month_cell.hyperlink = f"#'{target_sheet}'!A1"

        apply_data(ws, row, 2, mname_en, bg=bg, align=left_al())

        # SUMIFS formula: sum all transactions for this month from transactions sheet
        sumif_formula = (
            f"=SUMIF('{txns_sheet}'!$F:$F,\"{mname_he}\",'{txns_sheet}'!$D:$D)"
        )
        c_usd = ws.cell(row=row, column=3, value=sumif_formula)
        c_usd.font = Font(bold=True, size=11, name="Arial")
        c_usd.fill = fill(bg)
        c_usd.border = border_thin()
        c_usd.number_format = '[$$-en-US]#,##0.00'
        c_usd.alignment = right()

        c_ils = ws.cell(row=row, column=4, value=f"=SUMIF('{txns_sheet}'!$F:$F,\"{mname_he}\",'{txns_sheet}'!$E:$E)")
        c_ils.font = Font(bold=True, size=11, name="Arial")
        c_ils.fill = fill(bg)
        c_ils.border = border_thin()
        c_ils.number_format = '#,##0.00" \u20aa"'
        c_ils.alignment = right()

        total_row_ref = 4 + len(MONTHS) + 1
        c_pct = ws.cell(row=row, column=5)
        c_pct.value = f"=IF(C{total_row_ref}>0,C{row}/C{total_row_ref},0)"
        c_pct.font = Font(size=10, name="Arial")
        c_pct.fill = fill(bg)
        c_pct.border = border_thin()
        c_pct.number_format = '0.0%'
        c_pct.alignment = center()

        ws.cell(row=row, column=6).fill = fill(bg)
        ws.cell(row=row, column=6).border = border_thin()

        ws.row_dimensions[row].height = 20
        row += 1

    # Separator
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill(C_LTBLUE)
        ws.cell(row=row, column=ci).border = border_thin()
    row += 1

    # Grand Total
    ws.merge_cells(f"A{row}:B{row}")
    tc = ws.cell(row=row, column=1, value='\u05e1\u05d4"\u05db \u05e9\u05e0\u05ea\u05d9')
    tc.font = Font(bold=True, color=C_WHITE, size=13, name="Arial")
    tc.fill = fill(C_NAVY)
    tc.alignment = center()
    tc.border = border_medium()
    ws.cell(row=row, column=2).fill = fill(C_NAVY)
    ws.cell(row=row, column=2).border = border_medium()

    gt_usd = ws.cell(row=row, column=3, value=f"=SUM(C4:C{row-2})")
    gt_usd.font = Font(bold=True, color=C_WHITE, size=13, name="Arial")
    gt_usd.fill = fill(C_NAVY)
    gt_usd.border = border_medium()
    gt_usd.number_format = '[$$-en-US]#,##0.00'
    gt_usd.alignment = right()

    gt_ils = ws.cell(row=row, column=4, value=f"=SUM(D4:D{row-2})")
    gt_ils.font = Font(bold=True, color=C_WHITE, size=13, name="Arial")
    gt_ils.fill = fill(C_NAVY)
    gt_ils.border = border_medium()
    gt_ils.number_format = '#,##0.00" \u20aa"'
    gt_ils.alignment = right()

    gt_pct = ws.cell(row=row, column=5, value="100%")
    gt_pct.font = Font(bold=True, color=C_WHITE, size=11, name="Arial")
    gt_pct.fill = fill(C_NAVY)
    gt_pct.border = border_medium()
    gt_pct.number_format = '0.0%'
    gt_pct.alignment = center()

    ws.cell(row=row, column=6).fill = fill(C_NAVY)
    ws.cell(row=row, column=6).border = border_medium()
    ws.row_dimensions[row].height = 28

    avg_row = row + 1
    ws.merge_cells(f"A{avg_row}:B{avg_row}")
    ws.cell(row=avg_row, column=1).value = "\u05de\u05de\u05d5\u05e6\u05e2 \u05d7\u05d5\u05d3\u05e9\u05d9"
    ws.cell(row=avg_row, column=1).font = Font(bold=True, size=10, color=C_DARKGRAY, name="Arial")
    ws.cell(row=avg_row, column=1).fill = fill(C_LTBLUE)
    ws.cell(row=avg_row, column=1).alignment = center()
    ws.cell(row=avg_row, column=1).border = border_thin()
    ws.cell(row=avg_row, column=2).fill = fill(C_LTBLUE)
    ws.cell(row=avg_row, column=2).border = border_thin()

    avg_usd = ws.cell(row=avg_row, column=3, value=f"=AVERAGE(C4:C{row-2})")
    avg_usd.font = Font(bold=True, size=10, color=C_DARKGRAY, name="Arial")
    avg_usd.fill = fill(C_LTBLUE)
    avg_usd.border = border_thin()
    avg_usd.number_format = '[$$-en-US]#,##0.00'
    avg_usd.alignment = right()

    avg_ils = ws.cell(row=avg_row, column=4, value=f"=AVERAGE(D4:D{row-2})")
    avg_ils.font = Font(bold=True, size=10, color=C_DARKGRAY, name="Arial")
    avg_ils.fill = fill(C_LTBLUE)
    avg_ils.border = border_thin()
    avg_ils.number_format = '#,##0.00" \u20aa"'
    avg_ils.alignment = right()

    for ci in [5, 6]:
        ws.cell(row=avg_row, column=ci).fill = fill(C_LTBLUE)
        ws.cell(row=avg_row, column=ci).border = border_thin()

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly AI & Business Tools Expenses (USD)"
    chart.y_axis.title = "USD ($)"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data_ref = Reference(ws, min_col=3, min_row=3, max_row=row - 2)
    cats = Reference(ws, min_col=1, min_row=4, max_row=row - 2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, f"A{avg_row + 2}")

    set_col_width(ws, 1, 20)
    set_col_width(ws, 2, 18)
    set_col_width(ws, 3, 14)
    set_col_width(ws, 4, 14)
    set_col_width(ws, 5, 12)
    set_col_width(ws, 6, 22)

    return ws


# ════════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]

    build_settings(wb)
    build_transactions(wb)
    month_sheet_names = build_month_sheets(wb)
    build_monthly_summary(wb, month_sheet_names)
    build_annual_report(wb, month_sheet_names)

    for i, sheet in enumerate(wb.worksheets):
        if sheet.title == "\u05d3\u05d5\u05d7 \u05e9\u05e0\u05ea\u05d9":
            wb.active = i

    output_dir = os.path.dirname(OUTPUT_PATH)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    generate_dashboard_json()


def generate_dashboard_json():
    """Generate data.json for GitHub Pages dashboard."""
    import json
    from pathlib import Path
    from collections import defaultdict
    import datetime as dt

    docs_dir = (
        Path("docs")
        if os.environ.get("CI")
        else Path(r"C:\Users\roita\מעקב הוצאות כלים\docs")
    )
    docs_dir.mkdir(exist_ok=True)

    today = dt.date.today()
    dashboard_transactions = [
        transaction for transaction in TRANSACTION_RECORDS
        if dt.date.fromisoformat(transaction["date"]) <= today
    ]

    txns = [
        {
            "date": transaction["date"],
            "tool": transaction["tool"],
            "description": transaction["description"],
            "currency": transaction["currency"],
            "original_amount": transaction["original_amount"],
            "amount_usd": round(transaction["amount_usd"], 6),
            "amount_ils": transaction["amount_ils"],
            "entry_source": transaction.get("entry_source"),
        }
        for transaction in dashboard_transactions
    ]
    monthly = defaultdict(float)
    monthly_ils = defaultdict(float)
    by_tool = defaultdict(float)
    by_tool_ils = defaultdict(float)
    for transaction in dashboard_transactions:
        ym = transaction["month_key"]
        monthly[ym] = round(monthly[ym] + transaction["amount_usd"], 6)
        monthly_ils[ym] = round(monthly_ils[ym] + transaction["amount_ils"], 2)
        by_tool[transaction["tool"]] = round(by_tool[transaction["tool"]] + transaction["amount_usd"], 6)
        by_tool_ils[transaction["tool"]] = round(by_tool_ils[transaction["tool"]] + transaction["amount_ils"], 2)

    import datetime as _dt
    usd_rate, last_update, source = get_current_usd_to_ils_rate()
    rate_fetched_at = _dt.datetime.now(_dt.timezone.utc).isoformat()
    grand_usd = round(sum(transaction["amount_usd"] for transaction in dashboard_transactions), 2)
    grand_ils = round(sum(transaction["amount_ils"] for transaction in dashboard_transactions), 2)
    cur_month  = today.strftime("%Y-%m")
    cur_usd    = round(monthly.get(cur_month, 0.0), 2)
    cur_ils    = round(monthly_ils.get(cur_month, 0.0), 2)

    # Read auto_invoice run status if available
    scanner_status = None
    status_path = Path(__file__).resolve().parent / "auto_invoice_status.json"
    if status_path.exists():
        try:
            scanner_status = json.loads(status_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    data = {
        "generated":              today.isoformat(),
        "built_at":               rate_fetched_at,
        "usd_rate":               usd_rate,
        "exchange_rate_fetched_at": rate_fetched_at,
        "exchange_rate_updated_at": last_update,
        "exchange_rate_source":   source,
        "grand_total":            grand_usd,
        "grand_total_ils":        grand_ils,
        "current_month":          cur_month,
        "current_month_total":    cur_usd,
        "current_month_total_ils": cur_ils,
        "scanner_status":         scanner_status,
        "transactions": sorted(txns, key=lambda x: x["date"], reverse=True),
        "monthly": dict(sorted((key, round(value, 2)) for key, value in monthly.items())),
        "monthly_ils": dict(sorted(monthly_ils.items())),
        "by_tool": dict(sorted(((key, round(value, 2)) for key, value in by_tool.items()), key=lambda x: -x[1])),
        "by_tool_ils": dict(sorted(by_tool_ils.items(), key=lambda x: -x[1]))
    }

    json_content = json.dumps(data, ensure_ascii=False, indent=2)
    (docs_dir / "data.json").write_text(json_content, encoding="utf-8")

    # Keep dashboard-web/public/data.json in sync
    dashboard_public = Path(__file__).resolve().parent / "dashboard-web" / "public" / "data.json"
    if dashboard_public.parent.exists():
        dashboard_public.write_text(json_content, encoding="utf-8")

    print(f"Generated docs/data.json (grand total: ${data['grand_total']} / ₪{data['grand_total_ils']})")


if __name__ == "__main__":
    main()
